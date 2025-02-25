# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

from openpyxl import load_workbook

with open("D:\\ACMG_report\\20241210_ACMG_report\\Carrier_screening_W402-01.txt") as test:
    t = test.read()
    genes = [i for i in t.split("\t")[1].split(',')]

wb = load_workbook("D:\\ACMG_report\\20241210_ACMG_report\\JB24_325_Carrier\\基因表格修改raw data\\Carrier_fullgenelist.xlsx", data_only=True)

sheet = wb.active

def check_number(): 
    original = []
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if ',' in cell.value:
                check = cell.value.split(',')
            else:
                check = cell.value
            if isinstance(check, str):
                original.append(check)
            elif isinstance(check, list):
                for i in check:
                    original.append(i)
    t = list(set(original))
    return len(t)
#check_number()

with open("D:\\ACMG_report\\20241210_ACMG_report\\Carrirer_modification_record_genes.txt","w") as record:
    rows_delete = []
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):  # Start from row 2 to skip the header
            for cell in row:
                if ',' in cell.value:
                    check = cell.value.split(',')
                else:
                    check = cell.value
                first_column_value = sheet.cell(row=cell.row, column=1).value
                if isinstance(check, str):
                    if check not in genes:
                        record.write(f"Row {cell.row}, gene {check},{first_column_value} is deleted! \n")
                        rows_delete.append(cell.row)
                elif isinstance(check, list):
                    check_list = []
                    for i in check_list:
                        if i in genes:
                            check_list.append(True)
                        else:
                            check_list.append(False)        
                    if not any(check_list):
                        record.write(f"Row {cell.row}, gene {check},{first_column_value} is deleted! \n")
                        rows_delete.append(cell.row)
                    elif any(check_list) in genes:
                        preserve =  [i for i,val in enumerate(check_list) if val]
                        cell.value = ','.join(list(map(check.__getitem__,preserve)))
                        record.write(f"Row {cell.row}, gene [{check}],is edited to [{cell.value}]! \n")
                        
                    
    for row_index in sorted(rows_delete, reverse=True):  # Delete rows from bottom to top
        sheet.delete_rows(row_index)       


wb.save("D:\\ACMG_report\\20241210_ACMG_report\\Carrier_checked_v3.xlsx")

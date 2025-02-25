# -*- coding: utf-8 -*-
"""
Created on Tue Oct  8 17:04:07 2024

@author: yr.chen
"""

import googletrans
import openpyxl

excel = openpyxl.load_workbook('D:\\Others\\ND_genelist\\ND_disease-gene_check_final.xlsx')

sheet = excel.active

pt = sheet['G']

translations = []

for cell in pt:
    translator = googletrans.Translator()
    translation = translator.translate(cell.value, dest='zh-tw') # 翻譯成繁體中文
    translations.append(translation.text)
    
for i, translation in enumerate(translations):
    sheet.cell(row=i+1, column=10).value = translation
    print(sheet.cell(row=i+1, column=1).value, "=" , translation)
    
excel.save('D:\\Others\\ND_genelist\\ND_disease-gene_check_final.xlsx')
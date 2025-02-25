# -*- coding: utf-8 -*-
"""
Created on Sat Feb  8 10:26:44 2025

@author: yr.chen
"""

import argparse
import subprocess
from glob import glob
import os
from concurrent.futures import ThreadPoolExecutor
import ast
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
import time

parser = argparse.ArgumentParser(description="Search for gene-disease association",formatter_class=argparse.RawTextHelpFormatter)
parser.add_argument('-k','--keyword', action='append', help= "Keywords for searching, can be used for multiple times.", type=str)
parser.add_argument('-e','--exclude',action='append', help="Optional. Exclude data with certain words, can be used for multiple times.", nargs='?',type=str)
parser.add_argument('-s', '--star', help="Confidence score (float or integer) for association filtering, >= 3.0 is recommended for strong association.",type=str)
parser.add_argument('-l', '--level', help="1 : only look up in the textmining file; \n2 : look up in the knowledge file and then check in the textmining file ", choices=['1','2'])
parser.add_argument('-p','--prefix', help="Prefix for output files",type = str)

args = parser.parse_args()

keys = args.keyword
exc = args.exclude
prefix = args.prefix.strip()
star = args.star.strip()
level = args.level.strip()

grep_input = [f'grep -i {i}' for i in keys]
if exc:
    grep_filter = f"grep {' '.join([f'-vie {i}' for i in exc])}"
    grep_input.append(grep_filter)
    

def lookup(lookupfile, level1=True, checkinput=None):
    
    if level1:
        print(f"\nDISEASES metadata : \n{lookupfile}" ) #file path
        if len(grep_input) > 1:    
            cmd = f"{grep_input[0]} {lookupfile}|"+'|'.join(grep_input[1:])+r"|awk -F '\\t' '$6 >= %s'" %(star)
        elif len(grep_input) == 1:
            cmd = f"{grep_input[0]} {lookupfile}|"+r"awk -F '\\t' '$6 >= %s' " %(star)
        else:
            print("No keywords as input! Program stops.")
            exit()
    else:
        print(f"\nDISEASES metadata : \n{lookupfile}\n{checkinput}\n" ) 
        cmd2 = r"""|cut -f2|sort -u|xargs -I {} sh -c 'grep -w {} %s|head -n 1|awk -F "\\t" "\$6 >= %s"' """ %(checkinput,star)
        if len(grep_input) > 1:
            cmd1 = f"{grep_input[0]} {lookupfile}|"+'|'.join(grep_input[1:])
        elif len(grep_input) == 1:
            cmd1 = f"{grep_input[0]} {lookupfile}"
        else:
            print("No keywords as input! Program stops.")
            exit()
        cmd = cmd1+cmd2
    print(f"\nCommand to execute:\n{cmd}")
    result = subprocess.run(cmd, shell=True, check=True,universal_newlines=True, stdout=subprocess.PIPE,stderr=subprocess.PIPE)
    
    return(result.stdout)
        

def DISEASES_check():
    
    if level == "1":
        lookupfile = "textmining_full.sorted.tsv"
        association = lookup(lookupfile)
    elif level == "2":
        lookupfile = "knowledge_full.sorted.tsv"
        association = lookup(lookupfile,level1 = False, checkinput="textmining_full.sorted.tsv")
    
    return(association)

def HGNC_check(gene_ori):
    HGNC_info_file = "HGNC_genes_20250115.info"
    cmd = r"""awk -v gene=%s '{ if ($2 == gene) { print "true"; found=1; exit } } END { if (!found) print "false" }' %s;done""" %(gene_ori,HGNC_info_file)
    HGNC_result = subprocess.run(cmd, shell=True, check=True,universal_newlines=True, stdout=subprocess.PIPE,stderr=subprocess.PIPE)
        
    if HGNC_result.stdout.strip() == "true":
        return(gene_ori)
    else:
        cmd_HGNC = r"""grep -E '[^[:alpha:]]%s[^[:alpha:]]' %s|awk -F'\t' '$11 ~ /%s/'|cut -f2|paste -s -d '\t' """ %(gene_ori,HGNC_info_file,gene_ori)
        print(f"Command for looking up {gene_ori} HGNC name: \n{cmd_HGNC}")
        name_HGNC = subprocess.run(cmd_HGNC, shell=True, universal_newlines=True, stdout=subprocess.PIPE,stderr=subprocess.PIPE)
        if name_HGNC.stdout.strip():
            name = name_HGNC.stdout.strip().split('\t')
            if len(name) == 1:
                gene = name[0]
                return(gene)
            else:
                gene = input(f"The HGNC candidate name for {gene_ori} is {name},key in the right one : ").strip()
                return(gene)
        else:
            print(f"\nWarning! Can not find the HGNC symbol for gene {gene_ori}")
            keyin = input(f"\tKey in HGNC name for {gene_ori}, if skippable, press ENTER:")
            if keyin is not None:
                gene = keyin.strip()
                return(gene)
            else:
                return(False)

association = DISEASES_check()

def data_dict():  
    Dict = {}
    HGNC_genelist = open(f"{prefix}_genes_{star}star.HGNC","w+")
    rmlist =[]
    for i in association.splitlines():
        line = i.split('\t')
        gene_ori = line[1]
        if gene_ori not in Dict:
            print(f"Checking {gene_ori} HGNC name...")
            HGNC_name = HGNC_check(gene_ori)
            if HGNC_name:
                disease,zscore,confidence,weblink = line[3],line[4],line[5],line[6]
                HGNC_genelist.write(f"{HGNC_name}\n")
                if HGNC_name == gene_ori: 
                    Dict[HGNC_name] = ['.',disease,zscore,confidence,weblink]
                else:
                    Dict[HGNC_name] = [gene_ori,disease,zscore,confidence,weblink]
                    Dict[gene_ori] = ''
                    rmlist.append(gene_ori)
            else:
                Dict[gene_ori] = ''
                rmlist.append(gene_ori)
        else:
            continue
    HGNC_genelist.close()
    for i in rmlist:
        del Dict[i]
    return(Dict)
    

def get_latest_file(directory, file_pattern="*clinvar*.txt"):
    # Use glob to match files based on the file pattern
    files = glob(os.path.join(directory, file_pattern))
    
    if not files:  # If no files are found, return None
        print("CANNOT find any clinvar file !!")
        exit(1)
    
    # Get the file with the latest modification time
    latest_file = max(files, key=os.path.getmtime)
    return latest_file         

latest_clinvar = get_latest_file(directory="/Database")
star_py = 'clinvar_2_star.py'

def clinvar_search(gene):
    cmd1 = r'''awk -v g="%s" -F '\t' '{if ($5 == g ) print}' %s |grep -i pathogenic|grep -vi conflict''' %(gene, latest_clinvar)
    cmd2 = r'''|python3 %s -p %s '''%(star_py,prefix)
    cmd = cmd1+cmd2
    print(f"\nCommand for cliinvar search:\n {cmd}")
    result = subprocess.run(cmd, shell=True, check=True,universal_newlines=True, stdout=subprocess.PIPE,stderr=subprocess.PIPE)
    try : 
        D = ast.literal_eval(result.stdout)
    except:
        print(f"{gene} has no clinvar output after filteing")
        return False
    if type(D) is dict:
        print(f"Processing {gene} clinvar filtering output...")
        return D

def organize(gene):
    output = []
    output.append(gene) #0
    disease_collection = []
    clinvarout = clinvar_search(gene)
    
    if clinvarout:
        for i in clinvarout['lines']:
            diseases = list(set(i.split("\t")[5].strip().split("|")))
            disease_collection.extend(j for j in diseases if j not in disease_collection)
        pipeout = '|'.join(disease_collection)    
        if all(k.lower() in pipeout.lower() for k in keys):
            output.append('V') #1
        else:
            output.append('X') #1
        if clinvarout['clin-star'] == '2':
            output.append("2+") #2
        else:
            output.append("1=") #2
    else:
        output.append("X") #1
        output.append("--") #2
        pipeout = False
    
    diseaseout= Dict[gene]
    output.extend(diseaseout[:-1]) #3-#6
    
    if pipeout:
        output.append(pipeout) #7
    else:
        output.append("--") #7
    
    output.append(diseaseout[-1]) #8
    return output
    
if __name__ == '__main__':
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    target_dir = os.path.join(script_dir,prefix)
    os.makedirs(target_dir, exist_ok=True)
    os.chdir(target_dir)
    
    current = time.strftime("%m%d%H%M%S",time.localtime())
    log = open(f"log_{current}",'a+')
    log.write(str(vars(args))+'\n')
    
    num_threads = 16
    Dict = data_dict()
    genes = [i for i in Dict.keys()]
    
    header = ['Gene.HGNC','clinvar_keyword','clinvar_level','Gene_original','DISEASE','Z-score','Confidence','clinvar_pathogenic','weblink']
    
    if genes:
        with ThreadPoolExecutor(max_workers=num_threads) as executor:
            results = executor.map(organize, genes)
        data = list(results)
        
        sorted_data = sorted(data, key = lambda x : x[5],reverse=True)
        wb = Workbook()
        ws = wb.active
        ws.append(header)  # Add headers

        # Write sorted data to the sheet
        i = 1
        for row in sorted_data:
            i = i+1
            ws.append(row[:-1])
            cell = ws[f"I{i}"]
            cell.hyperlink = row[-1]
            cell.value = row[-1]
        
        decimal_style = NamedStyle(name="decimal", number_format="0.000")
        text_style = NamedStyle(name="text", number_format="@")
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.style = text_style
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=7):
            for cell in row:
                cell.value = float(cell.value)
                cell.style = decimal_style
        
        # Save the xlsx file
        outputname = f"{prefix}_association_level{level}_star{star}"
        wb.save(f"{outputname}.xlsx")        
        print(f"Excel file saved as {os.path.join(os.getcwd(),outputname)}.xlsx")
        log.write(f"Excel file saved as {outputname}.xlsx"+'\n')
        
        # Save the xls file
        with open(f"{outputname}.xls", "w+") as file:
            file.write('\t'.join(header)+'\n')
            for row in sorted_data:
                line = '\t'.join(row)+'\n'
                file.write(line)
        print(f"Excel file saved as {os.path.join(os.getcwd(),outputname)}.xls")
        log.write(f"Excel file saved as {outputname}.xls"+'\n')
            
        print(f"Total gene number: {len(genes)}")
        print(f"Gene list:\n{','.join(genes)}")
        log.write(f"Total gene number: {len(genes)}"+'\n')
        log.close()
    else:
        print("Program stops since no gene was found with your keywords or confidence level")
        exit()
    

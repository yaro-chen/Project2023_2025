# -*- coding: utf-8 -*-
"""
Created on Fri Dec 20 13:12:39 2024

@author: yr.chen
"""

from openpyxl import load_workbook
from datetime import datetime
import re
wb = load_workbook("", data_only=True) # import excel file

sheet = wb.get_sheet_by_name('Information')

for row in sheet.iter_rows(min_row=2, max_col=21, max_row=123):
        try:
            column_address = row[19]
            address_all = row[19].value
            address = row[19].value.split(';')[0]
            print(address)
            column_version = row[4]
            version = row[4].value
            version_num = float(re.search(r'\d+\.\d+', version).group())
            print(version, version_num)
            column_date = row[5]
            date = row[5].value
            print(date)
            column_report = row[10]
            report_all = column_report.value
            report = report_all.split(';')[0]
            output = row[0]
            if address and report:
                column_address.value = address_all.replace("old","new",1)
                print("new:",column_address.value)
                column_version.value = 'V{:.2}'.format(version_num+0.1)
                print("new:", column_version.value)
                column_date.value = datetime.today().strftime('%Y.%m.%d')
                print("new", column_date.value)
                output.value = "Y;;"
            print("---------")
        except:
            continue
        
wb.save("C:\\Users\\yr.chen\\Downloads\\桌面\\new.xlsx")
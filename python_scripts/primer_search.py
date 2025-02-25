# -*- coding: utf-8 -*-
"""
Created on Fri Jan 24 16:04:10 2025

@author: yr.chen
"""

from Bio import SeqIO
from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord
from multiprocessing import Process, Queue, cpu_count
import os
import openpyxl
import gc


def hamming_distance(s1, s2):
    """
    Calculate the Hamming distance between two strings.
    """
    if len(s1) != len(s2):
        raise ValueError("Strings must be of the same length.")
    return sum(ch1 != ch2 for ch1, ch2 in zip(s1, s2))


def find_matches_with_mismatch(sequence, target, max_mismatches=2):
    """
    Search for full matches or matches with up to max_mismatches in a given sequence.
    """
    matches = []
    target_len = len(target)
    for i in range(len(sequence) - target_len + 1):
        substring = sequence[i:i + target_len]
        if hamming_distance(substring, target) <= max_mismatches:
            matches.append((i, substring))
    return matches


def find_combined_matches(sequence, target1, target2, max_mismatches=2, max_distance=2000):
    """
    Find occurrences of two target sequences in the input sequence with constraints.
    """
    matches1 = find_matches_with_mismatch(sequence, target1, max_mismatches)
    if matches1:
        print(f"Forawrd sequence: {target1} is found")
        matches2 = find_matches_with_mismatch(sequence, str(Seq(target2).reverse_complement()), max_mismatches)
    else:
        matches2 = find_matches_with_mismatch(sequence, target2, max_mismatches)
        if matches2:
           print(f"Reverse sequence: {target2} is found")
           matches1 = find_matches_with_mismatch(sequence, str(Seq(target1).reverse_complement()), max_mismatches)
    combined_matches = []

    for idx1, match1 in matches1:
        for idx2, match2 in matches2:
            if 0 < idx2 - (idx1 + len(target1)) < max_distance:
                # Extract the substring between the matches (inclusive of matched regions)
                matched_substring = sequence[idx1:idx2 + len(target2)]
                combined_matches.append((idx1, match1, idx2, match2, matched_substring))
    return combined_matches


def process_record(record, target1, target2, max_mismatches, max_distance):
    """
    Process a single sequence record and find matches.
    """
    species_name = record.description
    sequence = str(record.seq)
    matches = find_combined_matches(sequence, target1, target2, max_mismatches, max_distance)
    results = []
    for idx1, match1, idx2, match2, matched_substring in matches:
        record_id = "{}_match_{}_{}".format(species_name, idx1, idx2)
        results.append(SeqRecord(
            Seq(matched_substring),
            id=record_id,
            description="Substring from {} to {}".format(idx1, idx2),
        ))
    return results


def worker(input_queue, output_queue, target1, target2, max_mismatches, max_distance):
    """
    Worker process to process records in parallel.
    """
    while True:
        record = input_queue.get()
        if record is None:  # Stop signal
            break
        results = process_record(record, target1, target2, max_mismatches, max_distance)
        output_queue.put(results)


def writer(output_queue, output_file):
    """
    Writer process to write matching records to the output file.
    """
    with open(output_file, "a") as output_handle:
        while True:
            results = output_queue.get()
            if results is None:  # Stop signal
                break
            SeqIO.write(results, output_handle, "fasta-2line")


def process_fasta_in_parallel(fasta_file, output_file, target1, target2, max_mismatches=2, max_distance=2000, num_workers=None):
    """
    Process a huge FASTA file in parallel while keeping memory usage low.
    """
    if num_workers is None:
        num_workers = max(1, cpu_count() - 1)
        print(num_workers)

    # Clear the output file before starting
    open(output_file, "w").close()

    # Queues for communication between processes
    input_queue = Queue(maxsize=0)
    output_queue = Queue(maxsize=0)

    # Start worker processes
    workers = []
    for _ in range(num_workers):
        p = Process(target=worker, args=(input_queue, output_queue, target1, target2, max_mismatches, max_distance))
        p.start()
        workers.append(p)

    # Start writer process
    writer_process = Process(target=writer, args=(output_queue, output_file))
    writer_process.start()

    # Read the FASTA file and feed records to the input queue
    for record in SeqIO.parse(fasta_file, "fasta"):
        input_queue.put(record)

    # Signal workers to stop
    for _ in workers:
        input_queue.put(None)

    # Wait for workers to finish
    for p in workers:
        p.join()

    # Signal writer to stop and wait for it to finish
    output_queue.put(None)
    writer_process.join()

    print(f"Finished processing the FASTA file. Results saved to {output_file}")


# Example usage
if __name__ == "__main__":
   max_distance = 2000
   max_mismatches = 2
   
   info = openpyxl.load_workbook("D:\\Others\\primer_search\\Primer_info.xlsx",data_only=True)
   sheet = info.active   
   info = []
   for row in sheet.iter_rows(min_row=2, max_col=(sheet.max_column), max_row=sheet.max_row): 
       if row :
           row_values = [cell.value for cell in row]
           info.append(row_values)
   print(info)
            
   for row in info:
       print(f"Row check:{row}")
       fasta_file = "D:\\Others\\primer_search\\S_pyogenes_complete.fna" 
       output_file = f"D:\\Others\\primer_search\\{row[0]}_{row[2]}.fasta"  
       print(f"Output: {output_file}")
       target1 = row[3]
       print(f"Target1:{target1}")
       target2 = row[5]
       print(f"Target2:{target2}")
       
       process_fasta_in_parallel(fasta_file, output_file, target1, target2, max_mismatches, max_distance=max_distance)
       gc.collect()
   #os.system('find ./ -size 0 -print -delete') 